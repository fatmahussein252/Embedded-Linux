
import openpyxl
from openpyxl import load_workbook
import random
import openpyxl.workbook

class employees:
    id = 0
    # This function takes the employee name, job and salary and generates
    #  random id for him/her and appends this data to the excel sheet.
    def add_employee(sht,wrkbook,name,job,salary):
        id = random.randint(0,500)
        data_ls=[name,id,job,salary]
        sht.append(data_ls)
        wrkbook.save('employees database.xlsx')
        print('----- operation done successfully -----') 

    # This function takes the employee name to print his/her data.
    def print_data(sht,name):
        for row in sht.iter_rows(2,sht.max_row,values_only=True):
             if row[0] == name:
                 print('-----------------------')
                 print('Name: ' + row[0] + '\nJob: ' + row[2] + '\nSalary: ' + row[3] )
                 print('-----------------------')
                 return
        print('-----------------------\nName Not found\n-----------------------')
    
    # This function takes the employee name to remove his/her data.
    def remove_employee(sht,wrkbook,name):
        # define counter to get the index of the required row to be deleted.
        i=2
        for row in sht.iter_rows(2,sht.max_row,values_only=True):
            if row[0] == name:
                sht.delete_rows(idx=i,amount=1)
                wrkbook.save('employees database.xlsx')
                print('----- operation done successfully -----')
                return
            else:
                i=i+1
    
        print('-----------------------\nName Not found\n-----------------------')
         
# This function opens the required excel sheet or create one.
def write_excel_file(path):
    # open the sheet or create it for running the code for the first time.
    try:
        wrkbook=load_workbook(path)
        sht=wrkbook.active
    except:
        wrkbook=openpyxl.Workbook()
        sht=wrkbook.active
        sht.cell(row=1,column=1).value='Name' 
        sht.cell(row=1,column=2).value='ID'
        sht.cell(row=1,column=3).value='job'
        sht.cell(row=1,column=4).value='salary' 
        wrkbook.save(path)
    return sht,wrkbook

# The intro point and main logic of the code
def main():
    # create excel sheet to store data.
    sht,wrkbook=write_excel_file('employees database.xlsx')
    # user operation choice.
    while(True):
        print('Enter number of operation from the above list:\n 1-Add new employee \n 2-Print employee data \n 3-remove employee from the system \n 4-terminate \n: ', end=" ")
        try:
            operation=int(input())
            if operation == 1:
                name = input ("Enter the employee Name: ")
                job = input ("Enter the employee job: ")
                salary = input ("Enter the employee salary: ")
                employees.add_employee(sht,wrkbook,name,job,salary)
            elif operation == 2:
                name = input("Enter the employee name: ")
                employees.print_data(sht,wrkbook,name)
            elif operation == 3:
                name = input("Enter the employee name: ")
                employees.remove_employee(sht,wrkbook,name)
            elif operation == 4:
                exit()
            else:
                print('-----------------------\nwrong input, enter interger number from the above list\n-----------------------')
        except ValueError:
            print('-----------------------\nwrong input, enter interger number from the above list\n-----------------------')

main()
