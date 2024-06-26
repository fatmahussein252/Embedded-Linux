import openpyxl

def write_excel_file(path,prototypes):
    # creat the excel file and sheet
    wrkbook=openpyxl.Workbook()
    sht=wrkbook.active
    # Add titles to used columns
    sht.cell(row=1,column=1).value='function prototype' 
    sht.cell(row=1,column=2).value='function ID' 
    # loop on columns and add vales
    i=2
    for element in prototypes:
        sht.cell(row=i,column=1).value=element
        sht.cell(row=i,column=2).value=i-1
        i=i+1

    # save the file
    wrkbook.save(path)
    wrkbook.close()


# open headers file
with open(r'session3\parse header file\headerfile.h') as headerfile:
    prototypes=headerfile.read()
    prototypes=prototypes.split(sep='\n') # make list of line of prototypes
    write_excel_file(r'session3\parse header file\headerfile.xlsx',prototypes)
        

